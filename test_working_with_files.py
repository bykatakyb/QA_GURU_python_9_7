import csv
import os
import zipfile
from io import TextIOWrapper

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

current_file = os.path.abspath(__file__)
current_dir = os.path.dirname(current_file)
path = os.path.join(current_dir, "temp")
resources = os.path.join(current_dir, "resources")
zip_path = os.path.join(resources, "archive.zip")


@pytest.fixture
def archive_files():
    if not os.path.exists(resources):
        os.mkdir(resources)
    file_dir = os.listdir(path)
    with zipfile.ZipFile(zip_path, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for file in file_dir:
            add_file = os.path.join(path, file)
            zf.write(add_file, file)
    # AUTO DELETING ARCHIVE AFTER TEST
    # yield
    # os.remove(zip_path)


def test_archive_creating(archive_files):
    assert zipfile.ZipFile("resources/archive.zip") != None


def test_files_opening(archive_files):
    with zipfile.ZipFile("resources/archive.zip") as zip_file:
        with zip_file.open("pdf.pdf") as pdf:
            reader = PdfReader(pdf)
            text = reader.pages[0].extract_text()
            assert "onetwothreefour" in text

        with zip_file.open("csv.csv", "r") as csv_file:
            reader = list(csv.reader(TextIOWrapper(csv_file)))
            assert "first" == reader[1][0]

        with zip_file.open("xlsx.xlsx") as xlsx:
            workbook = load_workbook(xlsx)
            sheet = workbook.active
            assert sheet.cell(row=4, column=1).value == "fourth"
